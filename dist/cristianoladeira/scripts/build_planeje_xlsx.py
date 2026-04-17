# Gera planeje-espaco-campos.xlsx — execute: python scripts/build_planeje_xlsx.py
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "planeje-espaco-campos.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Respostas"

headers = [
    "Data/hora",
    "Nome",
    "E-mail",
    "Telefone",
    "Por onde veio",
    "Haras/Empresa",
    "Baias",
    "Colaboradores",
    "Banheiro (sanitário + chuveiro)",
    "Cozinha",
    "Receptivo",
    "Segurança",
    "Limpeza periódica",
    "Demanda específica",
    "Identidade visual",
]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.freeze_panes = "A2"
for i in range(1, len(headers) + 1):
    letter = get_column_letter(i)
    w = min(32, max(14, len(headers[i - 1]) + 2))
    ws.column_dimensions[letter].width = w

ws2 = wb.create_sheet("Perguntas do formulário", 1)
ws2.cell(1, 1, "Nº")
ws2.cell(1, 2, "Pergunta")
h1 = ws2.cell(1, 1)
h2 = ws2.cell(1, 2)
h1.font = Font(bold=True)
h2.font = Font(bold=True)

questoes = [
    ("1", "Qual o seu nome?"),
    ("2", "Qual o seu e-mail para contato?"),
    ("3", "Qual o seu telefone para contato?"),
    ("4", "Por onde você veio? (LR EVENTOS / TODA ARTE MARKETING / INSTAGRAM / INDICAÇÃO / OUTROS)"),
    ("5", "Quantos Haras/Empresa participarão do seu espaço e quais serão?"),
    ("6", "Quantas baias você precisará? (número)"),
    ("7", "Pretende alojar quantos colaboradores? (número)"),
    ("8", "Pretende colocar banheiro com sanitário e chuveiro? (Sim / Não)"),
    ("9", "Pretende montar cozinha? (Sim / Não)"),
    ("10", "Pretende montar receptivo? (Sim / Não)"),
    ("11", "Precisa de segurança? (Sim / Não)"),
    ("12", "Precisa de limpeza periódica? (Sim / Não)"),
    ("13", "Alguma demanda específica?"),
    ("14", "Pretende fazer identidade visual? (Sim / Não)"),
]
for r, (n, q) in enumerate(questoes, 2):
    ws2.cell(r, 1, n)
    ws2.cell(r, 2, q)

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 78

wb.save(OUT)
print(f"Salvo: {OUT}")
